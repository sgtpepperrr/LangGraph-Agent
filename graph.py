import ast
import os
import re
from typing import List, Dict, Any, Optional

from dotenv import load_dotenv
from langchain_community.agent_toolkits import JiraToolkit
from langchain_community.utilities import JiraAPIWrapper
from langchain_deepseek import ChatDeepSeek
from langchain_openai import ChatOpenAI
from langgraph.prebuilt import create_react_agent
from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pydantic import BaseModel, Field, SecretStr
import matplotlib
import json
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import pymysql
from langchain_tavily import TavilySearch
import requests
from bs4 import BeautifulSoup
import subprocess
import platform
import psutil
from urllib.parse import quote
import time
import csv
# import datetime
from datetime import datetime, timedelta
from pathlib import Path

# 加载环境变量
load_dotenv(override=True)

# 导入高德地图工具
from amap_tools import AMAP_TOOLS

# ✅ 创建Tavily搜索工具
search_tool = TavilySearch(max_results=5, topic="general")

# ✅ 创建谷歌搜索工具（不需要API KEY）
class GoogleSearchSchema(BaseModel):
    query: str = Field(description="搜索查询关键词")
    num_results: int = Field(default=5, description="返回结果数量，默认5个")

@tool(args_schema=GoogleSearchSchema)
def google_search(query: str, num_results: int = 5) -> str:
    """
    使用谷歌搜索获取信息，不需要API KEY。
    适用于搜索最新信息、新闻、技术文档等。
    """
    try:
        # 构建搜索URL
        search_url = f"https://www.google.com/search?q={quote(query)}&num={num_results}"
        
        # 设置请求头模拟浏览器
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(search_url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取搜索结果
        results = []
        search_results = soup.find_all('div', class_='g')
        
        for i, result in enumerate(search_results[:num_results]):
            title_elem = result.find('h3')
            link_elem = result.find('a')
            snippet_elem = result.find('span', class_='aCOpRe')
            
            if title_elem and link_elem:
                title = title_elem.get_text()
                link = link_elem.get('href', '')
                snippet = snippet_elem.get_text() if snippet_elem else "无描述"
                
                results.append({
                    'title': title,
                    'link': link,
                    'snippet': snippet
                })
        
        if results:
            return json.dumps(results, ensure_ascii=False, indent=2)
        else:
            return "未找到相关搜索结果"
            
    except Exception as e:
        return f"搜索失败: {str(e)}"

# ✅ 创建网页内容抓取工具
class WebScrapingSchema(BaseModel):
    url: str = Field(description="要抓取的网页URL")
    extract_text: bool = Field(default=True, description="是否只提取文本内容")

@tool(args_schema=WebScrapingSchema)
def web_scraping(url: str, extract_text: bool = True) -> str:
    """
    抓取指定网页的内容。
    可以获取网页的文本内容或HTML源码。
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        response.encoding = response.apparent_encoding
        
        if extract_text:
            soup = BeautifulSoup(response.text, 'html.parser')
            # 移除脚本和样式元素
            for script in soup(["script", "style"]):
                script.decompose()
            
            # 获取文本内容
            text = soup.get_text()
            # 清理文本
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = '\n'.join(chunk for chunk in chunks if chunk)
            
            # 限制返回内容长度
            if len(text) > 3000:
                text = text[:3000] + "\n\n[内容已截断...]"
            
            return text
        else:
            return response.text[:3000] + "\n\n[HTML内容已截断...]" if len(response.text) > 3000 else response.text
            
    except Exception as e:
         return f"网页抓取失败: {str(e)}"

# ✅ 创建文件操作工具
class FileOperationSchema(BaseModel):
    operation: str = Field(description="操作类型: read, write, append, list, delete, exists")
    file_path: str = Field(description="文件路径")
    content: str = Field(default="", description="写入的内容（仅用于write和append操作）")

@tool(args_schema=FileOperationSchema)
def file_operations(operation: str, file_path: str, content: str = "") -> str:
    """
    执行文件操作，包括读取、写入、追加、列出目录、删除文件、检查文件是否存在。
    支持的操作类型：read, write, append, list, delete, exists
    """
    try:
        path = Path(file_path)
        
        if operation == "read":
            if path.exists() and path.is_file():
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    if len(content) > 2000:
                        content = content[:2000] + "\n\n[文件内容已截断...]"
                    return content
            else:
                return f"文件不存在: {file_path}"
                
        elif operation == "write":
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, 'w', encoding='utf-8') as f:
                f.write(content)
            return f"✅ 文件已写入: {file_path}"
            
        elif operation == "append":
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, 'a', encoding='utf-8') as f:
                f.write(content)
            return f"✅ 内容已追加到文件: {file_path}"
            
        elif operation == "list":
            if path.exists() and path.is_dir():
                items = []
                for item in path.iterdir():
                    item_type = "目录" if item.is_dir() else "文件"
                    size = item.stat().st_size if item.is_file() else "-"
                    items.append(f"{item_type}: {item.name} ({size} bytes)")
                return "\n".join(items[:50])  # 限制显示50个项目
            else:
                return f"目录不存在: {file_path}"
                
        elif operation == "delete":
            if path.exists():
                if path.is_file():
                    path.unlink()
                    return f"✅ 文件已删除: {file_path}"
                else:
                    return f"❌ 无法删除目录，请使用专门的目录删除工具"
            else:
                return f"文件不存在: {file_path}"
                
        elif operation == "exists":
            return f"文件存在: {path.exists()}"
            
        else:
            return f"不支持的操作类型: {operation}"
            
    except Exception as e:
        return f"文件操作失败: {str(e)}"

# ✅ 创建系统信息工具
@tool
def get_system_info() -> str:
    """
    获取当前系统的基本信息，包括操作系统、CPU、内存、磁盘使用情况等。
    """
    try:
        info = {
            "操作系统": platform.system(),
            "系统版本": platform.release(),
            "处理器": platform.processor(),
            "CPU核心数": psutil.cpu_count(logical=False),
            "逻辑CPU数": psutil.cpu_count(logical=True),
            "CPU使用率": f"{psutil.cpu_percent(interval=1):.1f}%",
            "内存总量": f"{psutil.virtual_memory().total / (1024**3):.1f} GB",
            "内存使用率": f"{psutil.virtual_memory().percent:.1f}%",
            "磁盘使用率": f"{psutil.disk_usage('/').percent:.1f}%",
            "当前时间": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        return json.dumps(info, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return f"获取系统信息失败: {str(e)}"

# ✅ 创建命令执行工具
class CommandSchema(BaseModel):
    command: str = Field(description="要执行的命令")
    timeout: int = Field(default=30, description="超时时间（秒）")

@tool(args_schema=CommandSchema)
def execute_command(command: str, timeout: int = 30) -> str:
    """
    执行系统命令并返回结果。
    注意：仅执行安全的命令，避免执行可能损害系统的命令。
    """
    try:
        # 安全检查：禁止执行危险命令
        dangerous_commands = ['rm -rf', 'del', 'format', 'fdisk', 'mkfs', 'dd if=', 'shutdown', 'reboot']
        if any(dangerous in command.lower() for dangerous in dangerous_commands):
            return "❌ 拒绝执行潜在危险命令"
        
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=timeout
        )
        
        output = result.stdout if result.stdout else result.stderr
        if len(output) > 2000:
            output = output[:2000] + "\n\n[输出已截断...]"
            
        return f"命令执行结果 (返回码: {result.returncode}):\n{output}"
        
    except subprocess.TimeoutExpired:
        return f"❌ 命令执行超时 ({timeout}秒)"
    except Exception as e:
        return f"❌ 命令执行失败: {str(e)}"

# ✅ 创建时间和日期工具
class DateTimeSchema(BaseModel):
    operation: str = Field(description="操作类型: now, format, calculate, timezone")
    date_string: str = Field(default="", description="日期字符串（用于格式化或计算）")
    format_string: str = Field(default="%Y-%m-%d %H:%M:%S", description="日期格式")
    days_offset: int = Field(default=0, description="天数偏移量（用于日期计算）")

@tool(args_schema=DateTimeSchema)
def datetime_operations(operation: str, date_string: str = "", format_string: str = "%Y-%m-%d %H:%M:%S", days_offset: int = 0) -> str:
    """
    执行日期时间相关操作，包括获取当前时间、格式化日期、日期计算等。
    """
    try:
        if operation == "now":
            return datetime.datetime.now().strftime(format_string)
            
        elif operation == "format":
            if date_string:
                # 尝试解析日期字符串
                dt = datetime.datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                return dt.strftime(format_string)
            else:
                return "请提供要格式化的日期字符串"
                
        elif operation == "calculate":
            base_date = datetime.datetime.now()
            if date_string:
                base_date = datetime.datetime.fromisoformat(date_string.replace('Z', '+00:00'))
            
            new_date = base_date + datetime.timedelta(days=days_offset)
            return new_date.strftime(format_string)
            
        elif operation == "timezone":
            import time
            return f"当前时区: {time.tzname[0]}, UTC偏移: {time.timezone // 3600} 小时"
            
        else:
            return f"不支持的操作类型: {operation}"
            
    except Exception as e:
        return f"日期时间操作失败: {str(e)}"

# ✅ 创建CSV数据处理工具
class CSVOperationSchema(BaseModel):
    operation: str = Field(description="操作类型: read, write, analyze")
    file_path: str = Field(description="CSV文件路径")
    data: str = Field(default="", description="CSV数据（JSON格式，用于写入操作）")

@tool(args_schema=CSVOperationSchema)
def csv_operations(operation: str, file_path: str, data: str = "") -> str:
    """
    执行CSV文件操作，包括读取、写入、基本分析。
    """
    try:
        if operation == "read":
            df = pd.read_csv(file_path)
            # 限制显示行数
            if len(df) > 10:
                preview = df.head(10)
                return f"CSV文件预览（前10行）：\n{preview.to_string()}\n\n总行数: {len(df)}, 总列数: {len(df.columns)}"
            else:
                return f"CSV文件内容：\n{df.to_string()}"
                
        elif operation == "write":
            if data:
                import json
                json_data = json.loads(data)
                df = pd.DataFrame(json_data)
                df.to_csv(file_path, index=False)
                return f"✅ 数据已写入CSV文件: {file_path}"
            else:
                return "请提供要写入的数据（JSON格式）"
                
        elif operation == "analyze":
            df = pd.read_csv(file_path)
            analysis = {
                "行数": len(df),
                "列数": len(df.columns),
                "列名": list(df.columns),
                "数据类型": df.dtypes.to_dict(),
                "缺失值": df.isnull().sum().to_dict(),
                "数值列统计": df.describe().to_dict() if len(df.select_dtypes(include=['number']).columns) > 0 else "无数值列"
            }
            return json.dumps(analysis, ensure_ascii=False, indent=2, default=str)
            
        else:
            return f"不支持的操作类型: {operation}"
            
    except Exception as e:
        return f"CSV操作失败: {str(e)}"

# ✅ 创建SQL查询工具
description = """
当用户需要进行数据库查询工作时，请调用该函数。
该函数用于在指定MySQL服务器上运行一段SQL代码，完成数据查询相关工作，
并且当前函数是使用pymsql连接MySQL数据库。
本函数只负责运行SQL代码并进行数据查询，若要进行数据提取，则使用另一个extract_data函数。
"""

# 定义结构化参数模型
class SQLQuerySchema(BaseModel):
    sql_query: str = Field(description=description)

# 封装为 LangGraph 工具
@tool(args_schema=SQLQuerySchema)
def sql_inter(sql_query: str) -> str:
    """
    当用户需要进行数据库查询工作时，请调用该函数。
    该函数用于在指定MySQL服务器上运行一段SQL代码，完成数据查询相关工作，
    并且当前函数是使用pymsql连接MySQL数据库。
    本函数只负责运行SQL代码并进行数据查询，若要进行数据提取，则使用另一个extract_data函数。
    :param sql_query: 字符串形式的SQL查询语句，用于执行对MySQL中telco_db数据库中各张表进行查询，并获得各表中的各类相关信息
    :return：sql_query在MySQL中的运行结果。   
    """
    # print("正在调用 sql_inter 工具运行 SQL 查询...")
    
    # 加载环境变量
    load_dotenv(override=True)
    host = os.getenv('HOST')
    user = os.getenv('USER')
    mysql_pw = os.getenv('MYSQL_PW')
    db = os.getenv('DB_NAME')
    port = os.getenv('PORT')
    
    # 创建连接
    connection = pymysql.connect(
        host=host,
        user=user,
        passwd=mysql_pw,
        db=db,
        port=int(port),
        charset='utf8'
    )
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            results = cursor.fetchall()
            # print("SQL 查询已成功执行，正在整理结果...")
    finally:
        connection.close()

    # 将结果以 JSON 字符串形式返回
    return json.dumps(results, ensure_ascii=False)

# ✅ 创建数据提取工具
# 定义结构化参数
class ExtractQuerySchema(BaseModel):
    sql_query: str = Field(description="用于从 MySQL 提取数据的 SQL 查询语句。")
    df_name: str = Field(description="指定用于保存结果的 pandas 变量名称（字符串形式）。")

# 注册为 Agent 工具
@tool(args_schema=ExtractQuerySchema)
def extract_data(sql_query: str, df_name: str) -> str:
    """
    用于在MySQL数据库中提取一张表到当前Python环境中，注意，本函数只负责数据表的提取，
    并不负责数据查询，若需要在MySQL中进行数据查询，请使用sql_inter函数。
    同时需要注意，编写外部函数的参数消息时，必须是满足json格式的字符串，
    :param sql_query: 字符串形式的SQL查询语句，用于提取MySQL中的某张表。
    :param df_name: 将MySQL数据库中提取的表格进行本地保存时的变量名，以字符串形式表示。
    :return：表格读取和保存结果
    """
    print("正在调用 extract_data 工具运行 SQL 查询...")
    
    load_dotenv(override=True)
    host = os.getenv('HOST')
    user = os.getenv('USER')
    mysql_pw = os.getenv('MYSQL_PW')
    db = os.getenv('DB_NAME')
    port = os.getenv('PORT')

    # 创建数据库连接
    connection = pymysql.connect(
        host=host,
        user=user,
        passwd=mysql_pw,
        db=db,
        port=int(port),
        charset='utf8'
    )

    try:
        # 执行 SQL 并保存为全局变量
        df = pd.read_sql(sql_query, connection)
        globals()[df_name] = df
        # print("数据成功提取并保存为全局变量：", df_name)
        return f"✅ 成功创建 pandas 对象 `{df_name}`，包含从 MySQL 提取的数据。"
    except Exception as e:
        return f"❌ 执行失败：{e}"
    finally:
        connection.close()

# ✅创建Python代码执行工具
# Python代码执行工具结构化参数说明
class PythonCodeInput(BaseModel):
    py_code: str = Field(description="一段合法的 Python 代码字符串，例如 '2 + 2' 或 'x = 3\\ny = x * 2'")

@tool(args_schema=PythonCodeInput)
def python_inter(py_code):
    """
    当用户需要编写Python程序并执行时，请调用该函数。
    该函数可以执行一段Python代码并返回最终结果，需要注意，本函数只能执行非绘图类的代码，若是绘图相关代码，则需要调用fig_inter函数运行。
    """    
    g = globals()
    try:
        # 尝试如果是表达式，则返回表达式运行结果
        return str(eval(py_code, g))
    # 若报错，则先测试是否是对相同变量重复赋值
    except Exception as e:
        global_vars_before = set(g.keys())
        try:            
            exec(py_code, g)
        except Exception as e:
            return f"代码执行时报错{e}"
        global_vars_after = set(g.keys())
        new_vars = global_vars_after - global_vars_before
        # 若存在新变量
        if new_vars:
            result = {var: g[var] for var in new_vars}
            # print("代码已顺利执行，正在进行结果梳理...")
            return str(result)
        else:
            # print("代码已顺利执行，正在进行结果梳理...")
            return "已经顺利执行代码"

# ✅ 创建绘图工具
# 绘图工具结构化参数说明
class FigCodeInput(BaseModel):
    py_code: str = Field(description="要执行的 Python 绘图代码，必须使用 matplotlib/seaborn 创建图像并赋值给变量")
    fname: str = Field(description="图像对象的变量名，例如 'fig'，用于从代码中提取并保存为图片")

@tool(args_schema=FigCodeInput)
def fig_inter(py_code: str, fname: str) -> str:
    """
    当用户需要使用 Python 进行可视化绘图任务时，请调用该函数。

    注意：
    1. 所有绘图代码必须创建一个图像对象，并将其赋值为指定变量名（例如 `fig`）。
    2. 必须使用 `fig = plt.figure()` 或 `fig = plt.subplots()`。
    3. 不要使用 `plt.show()`。
    4. 请确保代码最后调用 `fig.tight_layout()`。
    5. 所有绘图代码中，坐标轴标签（xlabel、ylabel）、标题（title）、图例（legend）等文本内容，必须使用英文描述。

    示例代码：
    fig = plt.figure(figsize=(10,6))
    plt.plot([1,2,3], [4,5,6])
    fig.tight_layout()
    """
    # print("正在调用fig_inter工具运行Python代码...")

    current_backend = matplotlib.get_backend()
    matplotlib.use('Agg')

    local_vars = {"plt": plt, "pd": pd, "sns": sns}
    
    # ✅ 设置图像保存路径（动态获取当前项目目录）
    current_dir = os.path.dirname(os.path.abspath(__file__))  # 获取当前脚本所在目录
    images_dir = os.path.join(current_dir, "images")
    os.makedirs(images_dir, exist_ok=True)  # ✅ 自动创建 images 文件夹（如不存在）
    try:
        g = globals()
        exec(py_code, g, local_vars)
        g.update(local_vars)

        fig = local_vars.get(fname, None)
        if fig:
            image_filename = f"{fname}.png"
            abs_path = os.path.join(images_dir, image_filename)  # ✅ 绝对路径
            rel_path = os.path.join("images", image_filename)    # ✅ 返回相对路径（给前端用）

            fig.savefig(abs_path, bbox_inches='tight')
            return f"✅ 图片已保存，路径为: {rel_path}"
        else:
            return "⚠️ 图像对象未找到，请确认变量名正确并为 matplotlib 图对象。"
    except Exception as e:
        return f"❌ 执行失败：{e}"
    finally:
        plt.close('all')
        matplotlib.use(current_backend)

# ✅ 创建jira工具
jira = JiraAPIWrapper()
jira_toolkit = JiraToolkit.from_jira_api_wrapper(jira)
jira_tools = jira_toolkit.get_tools()

# ✅ 创建周报日期工具
@tool
def get_weekly_dates() -> dict:
    """获取本周的开始和结束日期，以及上周的开始和结束日期。"""
    today = datetime.now().date()
    # 假设每周第一天是星期一
    start_of_this_week = today - timedelta(days=today.weekday())
    end_of_this_week = start_of_this_week + timedelta(days=4)
    start_of_last_week = start_of_this_week - timedelta(weeks=1)
    end_of_last_week = end_of_this_week - timedelta(weeks=1)

    return {
        "last_week_start": start_of_last_week.strftime("%Y/%m/%d"),
        "last_week_end": end_of_last_week.strftime("%Y/%m/%d"),
        "this_week_start": start_of_this_week.strftime("%Y/%m/%d"),
        "this_week_end": end_of_this_week.strftime("%Y/%m/%d"),
    }


def get_issue_data(raw_data_string: str) -> List[Dict]:
    """
    从Jira工具的原始JSON字符串中提取并整理问题数据。
    该函数现在会处理非标准的混合字符串和非JSON格式（如Python列表）。
    """
    if not raw_data_string or not raw_data_string.strip():
        return []

    # 1. 尝试匹配并提取 JSON/Python 列表部分
    # 匹配从 [ 开始到 ] 结束的部分
    match = re.search(r'\[\s*(?:{.*?}\s*,?\s*)*\]', raw_data_string, re.DOTALL)
    if not match:
        return []

    data_string  = match.group(0)

    try:
        # 2. 尝试使用 ast.literal_eval 来解析 Python 格式的列表
        tool_output_list = ast.literal_eval(data_string)
        if not isinstance(tool_output_list, list):
            return []
    except (ValueError, SyntaxError, TypeError):
        # 如果 ast.literal_eval 失败，可能是因为格式更接近JSON，
        # 或者其他不可解析的格式，直接返回空列表
        # print(f"解析工具返回的 JSON 时出错: {e}, 原始字符串: {data_string}")
        print(f"解析工具返回的 JSON 时出错, 原始字符串: {raw_data_string}")
        return []

    # 智能判断数据格式并提取 issues 列表
    issues = []
    if isinstance(tool_output_list, dict) and "issues" in tool_output_list:
        # 格式: {"issues": [{...}, {...}]}
        issues = tool_output_list.get("issues", [])
    elif isinstance(tool_output_list, list):
        # 格式: 直接就是issue列表 [{...}, {...}]
        issues = tool_output_list
    else:
        print(f"无法识别的JSON格式: {type(tool_output_list)}")
        return []

    extracted_issues = []
    for issue in issues:
        # 处理不同的数据结构
        if "fields" in issue:
            # 标准Jira格式: {"key": "PROJ-123", "fields": {"summary": "xxx"}}
            summary = issue.get("fields", {}).get("summary", "")
            key = issue.get("key", "")
        else:
            # 简化格式: 直接包含所有字段
            summary = issue.get("summary", "")
            key = issue.get("key", "")

        extracted_issues.append({
            "Project": key,
            "Task Description": summary,
            # 这里的其他字段留空，稍后在 excel 函数中填充
            "Estimated Start": "",
            "Estimated End": "",
            "Estimated Hours": "",
            "Actual Start": "",
            "Actual End": "",
            "Actual Hours": "",
            "Output": "",
            "Notes": ""
        })
    return extracted_issues

# ✅ 创建jira周报生成工具
def create_excel_report(done_issues, planned_issues, report_dates):
    """
    使用 openpyxl 库创建并保存周报 Excel 文件。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "周报"

    # 设置列头
    headers = [
        "Project", "Task Description", "预计开始", "预计结束", "预计时数",
        "实际开始", "实际结束", "实际工时", "产出", "备注"
    ]

    # 样式
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    section_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    row = 1
    # 写入上周已完成事项模块
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    cell = ws.cell(row=row, column=1, value="上周已完成事项")
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = section_fill

    row += 1
    # 写入列名
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header).font = header_font

    row += 1
    # 写入上周完成数据
    for issue in done_issues:
        ws.cell(row=row, column=1, value=issue.get("Project", "")).border = thin_border
        ws.cell(row=row, column=2, value=issue.get("Task Description", "")).border = thin_border
        ws.cell(row=row, column=3, value=report_dates["last_week_start"]).border = thin_border  # 预计开始
        ws.cell(row=row, column=4, value=report_dates["last_week_end"]).border = thin_border  # 预计结束
        ws.cell(row=row, column=5, value="40h").border = thin_border  # 预计时数
        ws.cell(row=row, column=6, value=report_dates["last_week_start"]).border = thin_border  # 实际开始
        ws.cell(row=row, column=7, value=report_dates["last_week_end"]).border = thin_border  # 实际结束
        ws.cell(row=row, column=8, value="40h").border = thin_border  # 实际工时
        ws.cell(row=row, column=9, value="").border = thin_border
        ws.cell(row=row, column=10, value="").border = thin_border
        row += 1

    row += 2  # 添加空行分隔

    # 写入本周预计进行事项模块
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    cell = ws.cell(row=row, column=1, value="本周预计进行事项")
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = section_fill

    row += 1
    # 写入列名
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header).font = header_font

    row += 1
    # 写入本周预计数据
    for issue in planned_issues:
        ws.cell(row=row, column=1, value=issue.get("Project", "")).border = thin_border
        ws.cell(row=row, column=2, value=issue.get("Task Description", "")).border = thin_border
        ws.cell(row=row, column=3, value=report_dates["this_week_start"]).border = thin_border
        ws.cell(row=row, column=4, value=report_dates["this_week_end"]).border = thin_border
        ws.cell(row=row, column=5, value="40h").border = thin_border
        ws.cell(row=row, column=6, value="").border = thin_border
        ws.cell(row=row, column=7, value="").border = thin_border
        ws.cell(row=row, column=8, value="").border = thin_border
        ws.cell(row=row, column=9, value="").border = thin_border
        ws.cell(row=row, column=10, value="").border = thin_border
        row += 1

    # 调整列宽
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = ws.cell(row=2, column=col_idx).column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    filename = f"周报_{report_dates['this_week_start'].replace('/', '-')}.xlsx"
    wb.save(filename)
    print(f"\n✅ 周报已成功生成并保存为：{filename}")

# 定义工具的参数 Schema
class CreateReportExcelSchema(BaseModel):
    raw_done_issues: str = Field(
        description="Jira 工具返回的原始上周已完成事项 JSON 字符串。"
    )
    raw_planned_issues: str = Field(
        description="Jira 工具返回的原始本周计划进行事项 JSON 字符串。"
    )
    report_dates: Dict[str, str] = Field(
        description="包含周报日期范围的字典。"
    )

@tool(args_schema=CreateReportExcelSchema)
def create_report_excel_tool(raw_done_issues: str, raw_planned_issues: str, report_dates: Dict) -> str:
    """
    根据Jira工具提供的原始JSON字符串，生成一份周报 Excel 文件。
    此工具负责数据解析、格式化并保存。
    """
    try:
        done_issues = get_issue_data(raw_done_issues)
        planned_issues = get_issue_data(raw_planned_issues)

        create_excel_report(done_issues, planned_issues, report_dates)
        filename = f"周报_{report_dates['this_week_start'].replace('/', '-')}.xlsx"
        return f"✅ 周报 Excel 文件已成功生成并保存为: {filename}"
    except Exception as e:
        return f"❌ 生成周报 Excel 失败: {str(e)}"

# ✅ 创建提示词模板
prompt = """
你是一名经验丰富的智能助手，擅长帮助用户高效完成各种任务。你拥有以下强大的工具能力：

## 🔍 搜索和信息获取工具
1. **Tavily搜索 (search_tool)**: 高质量的网络搜索，适合获取最新信息和新闻
2. **谷歌搜索 (google_search)**: 免费的谷歌搜索，无需API KEY，适合一般信息查询
3. **网页抓取 (web_scraping)**: 抓取指定网页的内容，可获取文本或HTML源码

## 📁 文件和系统操作工具
4. **文件操作 (file_operations)**: 读取、写入、追加、列出目录、删除文件、检查文件存在性
5. **系统信息 (get_system_info)**: 获取操作系统、CPU、内存、磁盘等系统信息
6. **命令执行 (execute_command)**: 安全地执行系统命令（禁止危险命令）
7. **时间日期 (datetime_operations)**: 获取当前时间、格式化日期、日期计算等
8. **CSV处理 (csv_operations)**: 读取、写入、分析CSV文件

## 💾 数据库和数据分析工具
9. **SQL查询 (sql_inter)**: 连接MySQL数据库执行查询，内置连接参数
10. **数据提取 (extract_data)**: 从MySQL提取数据到Python pandas环境
11. **Python执行 (python_inter)**: 执行非绘图类Python代码，数据处理和计算
12. **绘图工具 (fig_inter)**: 执行matplotlib/seaborn绘图代码并保存图片

## 🗺️ 高德地图API工具
13. **地理编码 (amap_geocode)**: 地址转换为经纬度坐标
14. **逆地理编码 (amap_regeocode)**: 坐标转换为详细地址信息
15. **POI搜索 (amap_poi_search)**: 搜索兴趣点，如餐厅、加油站、医院等
16. **输入提示 (amap_input_tips)**: 搜索建议和自动补全
17. **驾车路径规划 (amap_driving_route)**: 驾车导航路线规划
18. **步行路径规划 (amap_walking_route)**: 步行路线规划
19. **天气查询 (amap_weather)**: 实时天气和天气预报
20. **坐标转换 (amap_coord_convert)**: 不同坐标系之间的转换
21. **IP定位 (amap_ip_location)**: 根据IP地址获取位置
22. **行政区域查询 (amap_district_search)**: 查询省市区行政区域信息

## 🛠️ Jira 项目管理工具
23. **Jira 查询 (jql_query)**: 使用 JQL 语法来搜索和筛选 Jira 任务。
24. **Jira 获取项目列表 (get_projects)**: 获取所有你可访问的 Jira 项目列表。
25. **Jira 创建任务 (create_issue)**: 在指定的 Jira 项目中创建新的任务或缺陷。
26. **Jira 通用 API (catch_all_jira_api)**: 这是一个强大的通用工具，当现有工具无法满足你的需求时，你可以调用它来执行任何其他 Jira API 操作。

## 🎯 工具使用指南

**搜索信息时：**
- 最新新闻、实时信息 → 使用 `search_tool` 或 `google_search`
- 特定网页内容 → 使用 `web_scraping`

**文件操作时：**
- 读写文件 → 使用 `file_operations`
- 处理CSV数据 → 使用 `csv_operations`
- 系统信息查询 → 使用 `get_system_info`
- 执行命令 → 使用 `execute_command`（安全限制）

**数据分析时：**
- 数据库查询 → 先用 `sql_inter` 或 `extract_data`
- 数据处理 → 使用 `python_inter`
- 数据可视化 → 使用 `fig_inter`（必须创建fig对象，不要用plt.show()）

**时间处理时：**
- 获取当前时间、日期计算 → 使用 `datetime_operations`

**管理 Jira 任务时：**
- 查找任务: 使用 `jql_query`。你需要将用户的自然语言请求转化为对应的 JQL 语句。
    **重要**: 在你的 JQL 语句中，请使用 **`Executor` 字段**来指定当前用户，而不是 `assignee`，除非用户要求使用其它的字段。
- 创建任务: 使用 `create_issue`。
- 列出项目: 使用 `get_projects`。
- 执行复杂操作: 当用户需要执行的动作（如：更新任务、获取用户信息等）不被现有工具直接支持时，使用 `catch_all_jira_api`。这是一个强大的通用工具，你可以调用它来执行任何其他 Jira API 操作。
    输入格式: 一个字典，指定要调用的函数 (function)、参数 (args) 和关键字参数 (kwargs)。

**生成周报时：**
- **任务目标：** 自动生成包含上周已完成和本周计划事项的周报 Excel 文件。
- **步骤指引：**
    1. **获取日期：** **必须**使用 `get_weekly_dates` 工具来获取周报所需的精确日期范围。
    2. **查询数据：**
       - **上周已完成事项：** 使用你的 **Jira 搜索工具**，查询在 **`last_week_start` 到 `this_week_start`** 这个时间段内有**工时日志（Worklog）**记录的所有问题。
       - **重要**: 查询**上周已完成事项**的时间范围必须在**`last_week_start` 到 `this_week_start`**之间！！！
       - **本周计划事项：** 同样使用 **Jira 搜索工具**，查询当前用户**状态为 'Open' 或 'In Progress'** 的所有问题，并且不用过滤时间。
       - **重要**: 在你的 JQL 语句中，请使用 **`Executor` 字段**来指定当前用户，而不是 `assignee`，除非用户要求使用其它的字段。
    3. **生成文件：** 将上述 Jira 工具返回的**原始 JSON 字符串**和日期范围作为参数，调用 `create_report_excel_tool` 工具，以生成格式化好的周报文件。**注意：** `create_report_excel_tool` 内部会处理数据解析和格式化。

## ⚠️ 重要注意事项

1. **绘图要求：**
   - 必须使用 `fig = plt.figure()` 或 `fig, ax = plt.subplots()` 创建图像对象
   - 不要使用 `plt.show()`
   - 图表标签和标题使用英文
   - 调用 `fig.tight_layout()`

2. **安全限制：**
   - 命令执行工具禁止危险命令（如rm -rf, format等）
   - 文件操作限制在安全范围内

3. **输出格式：**
   - 使用**简体中文**回答
   - JSON数据要提取关键信息简要说明
   - 生成图片时使用Markdown格式：`![描述](images/图片名.png)`
   - 保持专业、简洁的风格

4. **工具选择优先级：**
   - 数据库操作 → SQL工具 → Python分析 → 绘图
   - 信息查询 → 搜索工具 → 网页抓取
   - 文件处理 → 专用文件工具 → 通用Python工具

请根据用户需求选择最合适的工具，提供准确、高效的帮助。
"""

# ✅ 创建工具列表
tools = [
    search_tool, 
    google_search,
    web_scraping,
    file_operations,
    get_system_info,
    execute_command,
    datetime_operations,
    csv_operations,
    python_inter, 
    fig_inter, 
    sql_inter, 
    extract_data,
    *jira_tools,  # 使用 * 解包操作符将 Jira 工具添加到列表中
    create_report_excel_tool,
    get_weekly_dates,
] + AMAP_TOOLS  # 添加高德地图工具

# ✅ 创建模型
# model = ChatDeepSeek(model="deepseek-chat")

# model = ChatOpenAI(
#     model="deepseek-r1:32b",
#     base_url="http://172.28.31.93:11434/v1/",
#     api_key="none",
#     # streaming=True
# )

model = ChatOpenAI(
    model="qwen-max-latest",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
# api_key=SecretStr(""),
    api_key=OPENAI_API_KEY,
)

# ✅ 创建图 （Agent）
graph = create_react_agent(model=model, tools=tools, prompt=prompt)
